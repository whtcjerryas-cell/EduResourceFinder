#!/usr/bin/env python3
"""
Excel导出器 - 生成详细的搜索日志分析报告
"""
import json
from typing import List, Dict, Any
from pathlib import Path
from logger_utils import get_logger

logger = get_logger('excel_exporter')


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self):
        """初始化Excel导出器"""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            self.Workbook = Workbook
            self.Font = Font
            self.Alignment = Alignment
            self.PatternFill = PatternFill
            self.Border = Border
            self.Side = Side
            
            self.openpyxl = openpyxl
            self.available = True
        except ImportError:
            logger.warning("openpyxl未安装，Excel导出功能不可用")
            self.available = False
    
    def export_search_log(self, search_log, output_path: str) -> bool:
        """
        导出搜索日志为Excel文件
        
        Args:
            search_log: SearchLog对象
            output_path: 输出文件路径
            
        Returns:
            是否成功
        """
        if not self.available:
            logger.error("openpyxl未安装，无法导出Excel")
            return False
        
        try:
            wb = self.Workbook()
            
            # Sheet 1: 模型输入输出表
            self._create_model_io_sheet(wb, search_log)
            
            # Sheet 2: 搜索结果表
            self._create_search_results_sheet(wb, search_log)
            
            # 保存文件
            wb.save(output_path)
            logger.info(f"✅ Excel文件已生成: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"导出Excel失败: {str(e)}")
            return False
    
    def _create_model_io_sheet(self, wb, search_log):
        """创建模型输入输出表（Sheet 1）"""
        ws = wb.active
        ws.title = "模型输入输出"
        
        # 定义表头
        headers = [
            "模型名称",
            "模型功能",
            "提供商",
            "时间戳",
            "提示词",
            "输入信息",
            "输出结果",
            "执行时间(秒)",
            "Token数",
            "成本"
        ]
        
        # 设置表头样式
        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = self.Font(bold=True, color="FFFFFF")
        header_alignment = self.Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入LLM调用数据
        for row_idx, llm_call in enumerate(search_log.llm_calls, 2):
            ws.cell(row=row_idx, column=1, value=llm_call.model_name)
            ws.cell(row=row_idx, column=2, value=llm_call.function)
            ws.cell(row=row_idx, column=3, value=llm_call.provider)
            ws.cell(row=row_idx, column=4, value=llm_call.timestamp)

            # 🔥 完整提示词 (启用换行)
            cell_prompt = ws.cell(row=row_idx, column=5, value=llm_call.prompt)
            cell_prompt.alignment = self.Alignment(wrap_text=True, vertical="top")

            # 🔥 完整输入信息 (启用换行)
            cell_input = ws.cell(row=row_idx, column=6, value=llm_call.input_data)
            cell_input.alignment = self.Alignment(wrap_text=True, vertical="top")

            # 🔥 完整输出 (启用换行)
            cell_output = ws.cell(row=row_idx, column=7, value=llm_call.output_data)
            cell_output.alignment = self.Alignment(wrap_text=True, vertical="top")

            ws.cell(row=row_idx, column=8, value=llm_call.execution_time)
            ws.cell(row=row_idx, column=9, value=llm_call.tokens_used or "")
            ws.cell(row=row_idx, column=10, value=llm_call.cost or "")
        
        # 自动调整列宽（不限制最大宽度，完整展示内容）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            # 🔥 不限制最大宽度，完整展示所有内容
            # 对于非常长的文本，Excel会自动处理
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        
        logger.info(f"✅ 模型输入输出表已创建: {len(search_log.llm_calls)}行")
    
    def _create_search_results_sheet(self, wb, search_log):
        """创建搜索结果表（Sheet 2）"""
        ws = wb.create_sheet(title="搜索结果")
        
        # 定义表头
        headers = [
            "搜索引擎",
            "查询关键词",
            "URL",
            "页面标题",
            "摘要",
            "评分",
            "推荐理由",
            "资源类型",
            "其他信息(JSON)"
        ]
        
        # 设置表头样式
        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = self.Font(bold=True, color="FFFFFF")
        header_alignment = self.Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入搜索结果数据
        for row_idx, result in enumerate(search_log.search_results, 2):
            ws.cell(row=row_idx, column=1, value=result.search_engine)
            ws.cell(row=row_idx, column=2, value=result.query)
            ws.cell(row=row_idx, column=3, value=result.url)
            ws.cell(row=row_idx, column=4, value=result.title)
            ws.cell(row=row_idx, column=5, value=result.snippet)
            ws.cell(row=row_idx, column=6, value=result.score)
            ws.cell(row=row_idx, column=7, value=result.recommendation_reason)
            ws.cell(row=row_idx, column=8, value=result.resource_type)
            
            # 其他信息转换为JSON字符串
            additional_info_json = json.dumps(result.additional_info, ensure_ascii=False, indent=2)
            ws.cell(row=row_idx, column=9, value=additional_info_json)
        
        # 自动调整列宽（不限制最大宽度，完整展示内容）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            # 🔥 不限制最大宽度，完整展示所有内容
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        
        # 设置文本换行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = self.Alignment(wrap_text=True, vertical="top")
        
        logger.info(f"✅ 搜索结果表已创建: {len(search_log.search_results)}行")
    
    def export_to_excel(self, search_log, output_path: str) -> bool:
        """
        导出为Excel文件
        
        Args:
            search_log: 搜索日志对象
            output_path: 输出路径
            
        Returns:
            是否成功
        """
        return self.export_search_log(search_log, output_path)
