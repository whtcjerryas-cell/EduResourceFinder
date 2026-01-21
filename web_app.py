#!/usr/bin/env python3
"""
Web 应用 - 教育视频搜索界面 V3
支持 AI 驱动的国家自动接入系统
"""

import os
import json
import time
import io
import sys
import uuid
import contextvars
import re
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Optional, Any

# ============================================================================
# 日志系统初始化（必须在导入其他模块之前）
# ============================================================================
from utils.logger_utils import get_logger
logger = get_logger('web_app')

# ============================================================================
# Request ID 上下文变量（用于关联日志）
# ============================================================================
request_id_var: contextvars.ContextVar[str] = contextvars.ContextVar('request_id', default='')

def get_request_id() -> str:
    """获取当前请求的 request_id"""
    return request_id_var.get('')

def set_request_id(request_id: str):
    """设置当前请求的 request_id"""
    request_id_var.set(request_id)

# 保存原始 print 函数
import builtins
_original_print = builtins.print

# 包装 print 函数，同时写入日志文件（包含 request_id）
def print(*args, **kwargs):
    """包装 print，同时写入日志文件，包含 request_id"""
    # 先调用原始 print（输出到控制台）
    _original_print(*args, **kwargs)
    # 同时写入日志文件
    message = ' '.join(str(arg) for arg in args)
    if message.strip():  # 只记录非空消息
        request_id = get_request_id()
        if request_id:
            message = f"[{request_id}] {message}"
        logger.info(message)

# ============================================================================
# Flask 应用初始化
# ============================================================================
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from flask_cors import CORS

app = Flask(__name__)

# ==============================================================================
# 注册蓝图模块（架构优化：拆分God Object）
# ==============================================================================
try:
    from routes import BLUEPRINT_CONFIG
    logger.info("🚀 开始注册蓝图模块...")

    for name, config in BLUEPRINT_CONFIG.items():
        try:
            init_func = config['init_func']
            url_prefix = config['url_prefix']
            bp = init_func()  # 初始化蓝图
            app.register_blueprint(bp, url_prefix=url_prefix)
            logger.info(f"  ✅ 已注册蓝图: {name} (前缀: {url_prefix or '/'})")
        except Exception as e:
            logger.error(f"  ❌ 蓝图 {name} 注册失败: {str(e)[:200]}")

    logger.info("✅ 蓝图注册完成")
except ImportError as e:
    logger.warning(f"⚠️ 蓝图模块导入失败: {str(e)}，将使用web_app.py中的路由")

# ==============================================================================
# 安全的 CORS 配置（修复：CORS Misconfiguration - P1 Critical）
# ==============================================================================
# 从环境变量读取允许的域名，默认为 localhost 开发环境
allowed_origins = os.getenv('ALLOWED_ORIGINS', 'http://localhost:3000').split(',')

# 严格的 CORS 配置：仅允许白名单域名访问 API
CORS(app, resources={
    r"/api/*": {
        "origins": allowed_origins,
        "methods": ["GET", "POST", "OPTIONS"],  # 仅允许必要的 HTTP 方法
        "allow_headers": ["Content-Type", "Authorization"],  # 仅允许必要的请求头
        "max_age": 3600,  # 预检请求缓存时间（秒）
        "supports_credentials": True  # 支持 credentials（如需要）
    }
})

logger.info(f"✅ CORS 已配置为仅允许以下来源: {allowed_origins}")

# 初始化并发限制器
try:
    from core.concurrency_limiter import get_concurrency_limiter
    concurrency_limiter = get_concurrency_limiter()
    logger.info("✅ 并发限制器已启用")
except ImportError:
    concurrency_limiter = None
    logger.warning("⚠️ 并发限制器未加载")

# 配置Flask日志
import logging
flask_logger = logging.getLogger('werkzeug')
flask_logger.setLevel(logging.WARNING)

# ============================================================================
# 配置和模块导入
# ============================================================================
from config_manager import ConfigManager
from core.grade_subject_validator import GradeSubjectValidator
from core.manual_review_system import ManualReviewSystem, ReviewStatus
from core.university_search_engine import UniversitySearchEngine, UniversitySearchRequest
from core.vocational_search_engine import VocationalSearchEngine, VocationalSearchRequest

# ✅ 安全修复：导入API密钥认证模块（Issue #041: Missing Authentication - FIXED）
from core.auth import require_api_key, require_admin, list_api_keys

config_manager = ConfigManager()
review_system = ManualReviewSystem()
university_search_engine = UniversitySearchEngine()
vocational_search_engine = VocationalSearchEngine()

# 尝试导入视频处理相关模块
try:
    from core.video_processor import VideoCrawler
    from core.video_evaluator import VideoEvaluator
    from core.playlist_processor import PlaylistProcessor
    HAS_VIDEO_PROCESSOR = True
    video_crawler = VideoCrawler()
    video_evaluator = VideoEvaluator()
    playlist_processor = PlaylistProcessor()
except ImportError as e:
    HAS_VIDEO_PROCESSOR = False
    video_crawler = None
    video_evaluator = None
    playlist_processor = None
    print(f"[⚠️ 警告] 视频处理模块不可用: {str(e)}")

# 导入简化的AI评估模块（不依赖视频下载）
try:
    from ai_evaluation import get_simple_evaluator
    HAS_AI_EVALUATION = True
    simple_evaluator = get_simple_evaluator()
    print(f"[✅ 成功] 简化AI评估模块已加载")
except ImportError as e:
    HAS_AI_EVALUATION = False
    simple_evaluator = None
    print(f"[⚠️ 警告] 简化AI评估模块不可用: {str(e)}")

# 搜索引擎将在每次请求时动态导入（避免模块缓存问题）
# 不在应用启动时导入，以确保每次都使用最新代码
HAS_SEARCH_ENGINE = True  # 假设模块可用，实际导入在请求时进行

# 尝试导入国家发现Agent
try:
    from tools.discovery_agent import CountryDiscoveryAgent
    HAS_DISCOVERY_AGENT = True
except ImportError as e:
    HAS_DISCOVERY_AGENT = False
    print(f"[⚠️ 警告] 国家发现模块不可用: {str(e)}")

# 导入服务类
try:
    from services.knowledge_overview_service import KnowledgeOverviewService
    from services.batch_video_service import BatchVideoService
    HAS_SERVICES = True
    knowledge_overview_service = KnowledgeOverviewService()
    print(f"[✅ 成功] 服务类已加载")
except ImportError as e:
    HAS_SERVICES = False
    knowledge_overview_service = None
    print(f"[⚠️ 警告] 服务类不可用: {str(e)}")

# ============================================================================
# 辅助函数
# ============================================================================

def _match_grade_to_knowledge_file(grade: str) -> str:
    """
    匹配年级到知识点文件名后缀
    
    Args:
        grade: 年级字符串（如 "Kelas 1", "1", "Kelas 4"）
    
    Returns:
        文件名后缀（如 "kelas1-2", "kelas3-4"），如果无法匹配则返回空字符串
    """
    grade_lower = grade.lower().strip()
    
    # 提取数字
    numbers = re.findall(r'\d+', grade_lower)
    
    if numbers:
        grade_num = int(numbers[0])
        # 映射到文件名后缀
        if grade_num <= 2:
            return "kelas1-2"
        elif grade_num <= 4:
            return "kelas3-4"
        elif grade_num <= 6:
            return "kelas5-6"
        elif grade_num <= 8:
            return "kelas7-8"
        elif grade_num <= 10:
            return "kelas9-10"
        elif grade_num <= 12:
            return "kelas11-12"
    
    # 如果没有数字，尝试字符串匹配
    if "kelas 1" in grade_lower or "kelas 2" in grade_lower or grade_lower in ["1", "2"]:
        return "kelas1-2"
    elif "kelas 3" in grade_lower or "kelas 4" in grade_lower or grade_lower in ["3", "4"]:
        return "kelas3-4"
    elif "kelas 5" in grade_lower or "kelas 6" in grade_lower or grade_lower in ["5", "6"]:
        return "kelas5-6"
    elif "kelas 7" in grade_lower or "kelas 8" in grade_lower or grade_lower in ["7", "8"]:
        return "kelas7-8"
    elif "kelas 9" in grade_lower or "kelas 10" in grade_lower or grade_lower in ["9", "10"]:
        return "kelas9-10"
    elif "kelas 11" in grade_lower or "kelas 12" in grade_lower or grade_lower in ["11", "12"]:
        return "kelas11-12"
    
    return ""

# ============================================================================
# 路由定义
# ============================================================================

@app.route('/')
def index():
    """根路由 - 返回前端页面"""
    return render_template('index.html')

@app.route('/favicon.ico')
def favicon():
    """返回favicon图标"""
    return send_from_directory('static', 'favicon.ico', mimetype='image/vnd.microsoft.icon')

@app.route('/app')
def full_app():
    """全教育层级搜索应用 - 支持K12/大学/职业三个层级"""
    return render_template('full_education_search.html')

@app.route('/knowledge_points')
def knowledge_points():
    """知识点概览页面"""
    return render_template('knowledge_points.html')

@app.route('/evaluation_reports')
def evaluation_reports():
    """评估报告页面"""
    return render_template('evaluation_reports.html')

@app.route('/search_history')
def search_history():
    """搜索历史页面"""
    return render_template('search_history.html')


@app.route('/test_base_new')
def test_base_new():
    """
    测试新的base_new.html模板
    用于验证UI统一方案的基础架构
    """
    return render_template('test_base_new.html')

@app.route('/test_search')
def test_search():
    """搜索API测试页面"""
    return send_from_directory('.', 'test_search.html')

@app.route('/api/countries', methods=['GET'])
def get_countries():
    """获取所有国家列表"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)
    
    try:
        countries = config_manager.get_all_countries()
        return jsonify({
            "success": True,
            "countries": countries
        })
    except Exception as e:
        logger.error(f"获取国家列表失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e),
            "countries": []
        }), 500

@app.route('/api/search_history', methods=['GET'])
def get_search_history():
    """获取搜索历史"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        history_file = os.path.join(os.path.dirname(__file__), 'search_history.json')

        if not os.path.exists(history_file):
            # 如果文件不存在，返回空列表
            return jsonify({
                "success": True,
                "history": [],
                "total": 0
            })

        with open(history_file, 'r', encoding='utf-8') as f:
            history = json.load(f)

        # 按时间倒序排列
        history = sorted(history, key=lambda x: x.get('timestamp', ''), reverse=True)

        return jsonify({
            "success": True,
            "history": history,
            "total": len(history)
        })

    except Exception as e:
        logger.error(f"获取搜索历史失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e),
            "history": [],
            "total": 0
        }), 500

@app.route('/api/config/<country_code>', methods=['GET'])
def get_config(country_code: str):
    """获取国家配置"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)
    
    try:
        config = config_manager.get_country_config(country_code.upper())
        if config:
            return jsonify({
                "success": True,
                "config": config.dict()
            })
        else:
            return jsonify({
                "success": False,
                "message": f"国家配置不存在: {country_code}",
                "config": None
            }), 404
    except Exception as e:
        logger.error(f"获取国家配置失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e),
            "config": None
        }), 500

@app.route('/api/config/education_levels', methods=['GET'])
def get_education_levels():
    """获取国家的教育层级（年级）"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        country = request.args.get('country', '').upper()
        if not country:
            return jsonify({
                "success": False,
                "message": "缺少country参数",
                "levels": []
            }), 400

        config = config_manager.get_country_config(country)
        if not config:
            return jsonify({
                "success": False,
                "message": f"国家配置不存在: {country}",
                "levels": []
            }), 404

        # 从 grade_subject_mappings 获取年级列表
        levels = []
        if hasattr(config, 'grade_subject_mappings') and config.grade_subject_mappings:
            levels = list(config.grade_subject_mappings.keys())

        return jsonify({
            "success": True,
            "levels": levels
        })
    except Exception as e:
        logger.error(f"获取教育层级失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e),
            "levels": []
        }), 500

@app.route('/api/config/subjects', methods=['GET'])
def get_subjects():
    """获取国家的学科列表"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        country = request.args.get('country', '').upper()
        if not country:
            return jsonify({
                "success": False,
                "message": "缺少country参数",
                "subjects": []
            }), 400

        config = config_manager.get_country_config(country)
        if not config:
            return jsonify({
                "success": False,
                "message": f"国家配置不存在: {country}",
                "subjects": []
            }), 404

        subjects = config.subjects if hasattr(config, 'subjects') else []
        return jsonify({
            "success": True,
            "subjects": subjects
        })
    except Exception as e:
        logger.error(f"获取学科列表失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e),
            "subjects": []
        }), 500

@app.route('/api/search', methods=['POST'])
@require_api_key  # ✅ 安全修复：需要API密钥认证
def search():
    """搜索API - 使用 Pydantic 验证输入（修复：输入验证缺失）"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    # 并发限制检查
    acquired_limiter = False
    if concurrency_limiter is not None:
        if concurrency_limiter.acquire(timeout=5.0):
            acquired_limiter = True
        else:
            logger.warning(f"搜索请求被限流: 超过最大并发数")
            return jsonify({
                "success": False,
                "message": "服务器繁忙，请稍后重试"
            }), 503

    try:
        logger.info(f"[搜索请求] 开始处理搜索请求 [ID: {request_id}]")
        logger.debug(f"[搜索请求] 请求数据: {json.dumps(request.get_json(), ensure_ascii=False)}")

        data = request.get_json()

        # ======================================================================
        # 输入验证（安全修复：P1 - 防止注入攻击）
        # ======================================================================
        from core.input_validators import validate_search_request

        is_valid, error_msg, validated_data = validate_search_request(data)
        if not is_valid:
            logger.warning(f"[搜索请求] 输入验证失败: {error_msg}")
            return jsonify({
                "success": False,
                "message": f"输入验证失败: {error_msg}",
                "results": []
            }), 400

        # 使用验证后的安全数据
        country = validated_data.country
        grade = validated_data.grade
        subject = validated_data.subject
        semester = validated_data.semester or None
        language = validated_data.language or None
        resource_type = validated_data.resource_type

        logger.info(f"[搜索参数] 国家={country}, 年级={grade}, 学科={subject}, 学期={semester}, 语言={language}, 资源类型={resource_type}")

        if not HAS_SEARCH_ENGINE:
            logger.error("[搜索请求] 搜索引擎模块不可用")
            return jsonify({
                "success": False,
                "message": "搜索引擎模块不可用",
                "results": []
            }), 500

        # 强制重新加载模块（确保获取最新代码）
        logger.debug("[搜索请求] 开始加载搜索引擎模块...")
        import importlib
        import search_engine_v2
        importlib.reload(search_engine_v2)
        from search_engine_v2 import SearchRequest, SearchEngineV2 as ReloadedSearchEngineV2
        logger.debug("[搜索请求] 搜索引擎模块加载完成")

        search_request = SearchRequest(
            country=country,
            grade=grade,
            semester=semester,
            subject=subject,
            language=language
        )
        
        logger.info(f"[搜索执行] 开始执行搜索 [ID: {request_id}]")

        # 📊 启动日志收集
        from core.search_log_collector import get_log_collector
        log_collector = get_log_collector()
        search_id = log_collector.start_search(country, grade, subject, semester)
        logger.info(f"[日志收集] 已启动搜索日志: {search_id}")

        import time
        import gc
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
        search_start_time = time.time()
        
        # 添加整体超时保护（200秒）- 使用ThreadPoolExecutor实现真正的超时中断
        SEARCH_TIMEOUT = 200  # 🔧 增加到200秒以支持LLM评估
        response = None
        search_engine_instance = None  # 用于内存清理
        
        def execute_search():
            """在独立线程中执行搜索"""
            nonlocal search_engine_instance
            # 传递 log_collector 给搜索引擎
            search_engine_instance = ReloadedSearchEngineV2(log_collector=log_collector)
            try:
                result = search_engine_instance.search(search_request)
                return result
            finally:
                # 在线程内部清理资源
                try:
                    if search_engine_instance is not None:
                        del search_engine_instance
                        gc.collect()
                except:
                    pass
        
        try:
            # 使用ThreadPoolExecutor执行搜索，支持真正的超时中断
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(execute_search)
                try:
                    response = future.result(timeout=SEARCH_TIMEOUT)
                    search_elapsed = time.time() - search_start_time
                    logger.info(f"[搜索执行] 搜索完成，耗时: {search_elapsed:.2f}秒，结果数: {len(response.results)}")
                except FuturesTimeoutError:
                    logger.error(f"[搜索执行] 搜索超时（超过{SEARCH_TIMEOUT}秒）[ID: {request_id}]")
                    # 尝试取消任务（虽然可能已经无法取消）
                    future.cancel()
                    # 返回超时响应
                    from search_engine_v2 import SearchResponse
                    response = SearchResponse(
                        success=False,
                        query="",
                        results=[],
                        message=f"搜索超时（超过{SEARCH_TIMEOUT}秒），请稍后重试或减少搜索条件",
                        total_count=0,
                        playlist_count=0,
                        video_count=0
                    )

            # 📊 记录搜索结果到日志
            search_elapsed = time.time() - search_start_time
            if response and response.success:
                for result in response.results:
                    # 🔥 获取真实的搜索引擎名称（从result对象中）
                    search_engine = getattr(result, 'search_engine', None) or (
                        result.model_dump().get('search_engine') if hasattr(result, 'model_dump') else None
                    ) or "Unknown"

                    log_collector.record_search_result(
                        engine=search_engine,  # 使用真实的搜索引擎名称
                        query=response.query,
                        url=result.url or "",
                        title=result.title or "",
                        snippet=result.snippet or "",
                        score=result.score or 0,
                        recommendation_reason=result.recommendation_reason or "",
                        resource_type=result.resource_type or "未知",
                    )
                # 完成日志收集
                log_collector.finish_search(
                    total_time=search_elapsed,
                    search_time=search_elapsed * 0.7,  # 估算搜索时间
                    scoring_time=search_elapsed * 0.3  # 估算评分时间
                )
                logger.info(f"[日志收集] 搜索日志已完成: {search_id}, 结果数: {len(response.results)}")
            else:
                logger.warning(f"[日志收集] 搜索失败，跳过日志记录: {search_id}")

        except Exception as e:
            search_error = str(e)
            logger.error(f"[搜索执行] 搜索异常: {search_error} [ID: {request_id}]")
            raise  # 重新抛出异常，让外层catch处理
        finally:
            # 🔥 搜索完成后立即清理内存，防止内存泄漏
            # 注意：search_engine_instance 在线程内部已经清理，这里只做额外清理
            try:
                import gc
                gc.collect()
                logger.debug("🗑️ 搜索完成，已清理内存")
            except Exception as e:
                logger.debug(f"内存清理: {str(e)}")

        # 保存搜索历史
        history_file = os.path.join(os.path.dirname(__file__), 'search_history.json')
        history = []
        if os.path.exists(history_file):
            try:
                with open(history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)
            except:
                history = []
        
        history.insert(0, {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "request": {
                "country": country,
                "grade": grade,
                "semester": semester,
                "subject": subject,
                "language": language
            },
            "response": {
                "success": response.success,
                "query": response.query,
                "total_count": response.total_count,
                "playlist_count": response.playlist_count,
                "video_count": response.video_count,
                "results": [r.model_dump() if hasattr(r, 'model_dump') else r.dict() for r in response.results]
            }
        })
        
        # 只保留最近100条
        history = history[:100]
        
        with open(history_file, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        
        # 获取最近的日志（用于前端Debug弹窗）
        debug_logs = []
        try:
            log_file = os.path.join(os.path.dirname(__file__), 'search_system.log')
            if os.path.exists(log_file):
                with open(log_file, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    debug_logs = lines[-100:] if len(lines) > 100 else lines
        except:
            pass

        # 根据资源类型过滤结果
        filtered_results = response.results

        # 统一清理标题（在返回前）
        from core.text_utils import clean_title
        for r in filtered_results:
            if r.title and '- YouTube' in r.title:
                r.title = clean_title(r.title, r.url or '')

        # ✅ 已移除URL过滤逻辑 - 显示所有搜索结果
        # 原有的should_exclude()函数已移除，不再过滤任何URL
        logger.info(f"[结果处理] 原始结果数: {len(filtered_results)}")

        if resource_type and resource_type != 'all':
            # 资源类型映射（'video'包括播放列表和视频）
            type_mapping = {
                'video': ['播放列表', '视频'],  # video包括播放列表和视频
                'textbook': '教材',
                'supplement': '教辅',
                'exercise': '练习题'
            }

            target_types = type_mapping.get(resource_type)

            if target_types:
                original_count = len(filtered_results)
                if isinstance(target_types, list):
                    # 如果是列表，匹配列表中的任一类型
                    filtered_results = [r for r in filtered_results if r.resource_type in target_types]
                    logger.info(f"[资源过滤] 类型={resource_type} ({', '.join(target_types)}), 原始={original_count}, 过滤后={len(filtered_results)}")
                else:
                    # 如果是单个类型，直接匹配
                    filtered_results = [r for r in filtered_results if r.resource_type == target_types]
                    logger.info(f"[资源过滤] 类型={resource_type} ({target_types}), 原始={original_count}, 过滤后={len(filtered_results)}")

        # ✅ 已移除播放列表优先和单个视频质量过滤 - 显示所有结果
        # 只按评分倒序排列（高分在前）

        # 统计信息
        def is_playlist(result):
            """判断是否是YouTube播放列表"""
            url = result.url or ''
            return bool(
                'playlist' in url.lower() or
                'list=' in url.lower() or
                '/videos' in url.lower() or
                result.resource_type == '播放列表'
            )

        playlist_count = sum(1 for r in filtered_results if is_playlist(r))
        logger.info(f"[结果统计] 总结果: {len(filtered_results)}, 播放列表: {playlist_count}, 单个视频: {len(filtered_results) - playlist_count}")

        # 按评分倒序排列（高分在前）
        filtered_results.sort(
            key=lambda r: -getattr(r, 'score', 0)  # 只按评分降序
        )

        logger.info(f"[结果排序] 按评分倒序，前5个结果:")
        for i, r in enumerate(filtered_results[:5], 1):
            score = getattr(r, 'score', 0)
            resource_type = getattr(r, 'resource_type', '未知')
            logger.info(f"  {i}. [{resource_type}] {score:.1f}/10 - {r.url[:60]}")

        # 记录最终结果统计
        logger.info(f"[搜索完成] 请求ID: {request_id}, 总结果数: {len(filtered_results)}, 播放列表: {response.playlist_count}, 视频: {response.video_count}")

        # 🔍 使用model_dump()获取所有字段（包括Optional字段）
        response_dict = response.model_dump()
        logger.info(f"[DEBUG] response.model_dump()字段: {list(response_dict.keys())}")
        logger.info(f"[DEBUG] 有quality_report: {'quality_report' in response_dict}")
        logger.info(f"[DEBUG] 有optimization_request: {'optimization_request' in response_dict}")

        # 构建响应数据
        response_data = {
            "success": response.success,
            "query": response.query,
            "results": [r.model_dump() if hasattr(r, 'model_dump') else r.dict() for r in filtered_results],
            "total_count": len(filtered_results),  # 使用过滤后的数量
            "playlist_count": response.playlist_count,
            "video_count": response.video_count,
            "message": response.message,
            "timestamp": response.timestamp,
            "debug_logs": debug_logs,
            "search_id": search_id  # 📊 添加搜索ID用于导出日志
        }

        # 添加质量评估报告（使用model_dump获取）
        if 'quality_report' in response_dict and response_dict['quality_report']:
            response_data["quality_report"] = response_dict['quality_report']
            logger.info(f"[质量评估] 质量分数: {response_dict['quality_report'].get('overall_quality_score', 'N/A')}/100")

        # 添加优化请求（使用model_dump获取）
        if 'optimization_request' in response_dict and response_dict['optimization_request']:
            response_data["optimization_request"] = response_dict['optimization_request']
            logger.info(f"[优化请求] 优化请求ID: {response_dict['optimization_request'].get('request_id', 'N/A')}")
            logger.info(f"[优化请求] 待审批方案数: {len(response_dict['optimization_request'].get('optimization_plans', []))}")

        # 🔍 添加透明度元数据（P0-1）
        if 'transparency' in response_dict and response_dict['transparency']:
            response_data["transparency"] = response_dict['transparency']
            transparency = response_dict['transparency']
            logger.info(f"[透明度] 搜索次数: {transparency.get('total_searches', 0)}, 原始结果: {transparency.get('total_raw_results', 0)}, 耗时: {transparency.get('total_duration_ms', 0)/1000:.1f}s")
        else:
            logger.debug("[透明度] 无透明度数据")

        logger.info(f"[DEBUG] 最终响应数据字段: {list(response_data.keys())}")

        return jsonify(response_data)
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        error_message = str(e)
        
        # 详细记录错误信息到日志
        logger.error(f"[搜索失败] 请求ID: {request_id}")
        logger.error(f"[搜索失败] 错误类型: {type(e).__name__}")
        logger.error(f"[搜索失败] 错误消息: {error_message}")
        logger.error(f"[搜索失败] 完整堆栈跟踪:\n{error_traceback}")
        
        # 同时打印到控制台（确保能看到）
        print(f"\n{'='*80}")
        print(f"❌ [搜索失败] 请求ID: {request_id}")
        print(f"❌ [搜索失败] 错误类型: {type(e).__name__}")
        print(f"❌ [搜索失败] 错误消息: {error_message}")
        print(f"❌ [搜索失败] 完整堆栈跟踪:")
        print(error_traceback)
        print(f"{'='*80}\n")
        
        # 检查是否是超时错误（错误代码5）
        if "timeout" in error_message.lower() or "超时" in error_message or "timed out" in error_message.lower():
            logger.error("[搜索失败] ⚠️ 检测到超时错误（错误代码5）")
            print("⚠️ 检测到超时错误（错误代码5）")
        
        traceback.print_exc()  # 也打印到stderr
        
        return jsonify({
            "success": False,
            "message": f"搜索失败: {error_message}",
            "error_type": type(e).__name__,
            "error_code": 5 if ("timeout" in error_message.lower() or "超时" in error_message) else None,
            "results": []
        }), 500
    finally:
        # 释放并发限制器（只有在成功获取许可的情况下才释放）
        if concurrency_limiter is not None and acquired_limiter:
            try:
                concurrency_limiter.release()
            except Exception as e:
                logger.error(f"释放并发限制器失败: {str(e)}")

@app.route('/api/history', methods=['GET'])
def get_history():
    """获取搜索历史"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)
    
    try:
        history_file = os.path.join(os.path.dirname(__file__), 'search_history.json')
        if os.path.exists(history_file):
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
            return jsonify({
                "success": True,
                "history": history
            })
        else:
            return jsonify({
                "success": True,
                "history": []
            })
    except Exception as e:
        logger.error(f"获取历史记录失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e),
            "history": []
        }), 500

@app.route('/api/evaluation_history', methods=['GET'])
def get_evaluation_history():
    """获取评估历史"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)
    
    try:
        evaluations_dir = os.path.join(os.path.dirname(__file__), 'data', 'evaluations')
        if not os.path.exists(evaluations_dir):
            return jsonify({
                "success": True,
                "evaluations": []
            })
        
        evaluations = []
        for filename in os.listdir(evaluations_dir):
            if filename.startswith('evaluation_') and filename.endswith('.json'):
                filepath = os.path.join(evaluations_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        eval_data = json.load(f)
                    # 检查是否是批量评估
                    is_batch = eval_data.get('is_batch', False)
                    total_videos = eval_data.get('total_videos', 0) if is_batch else None
                    
                    evaluations.append({
                        "request_id": eval_data.get('request_id', ''),
                        "timestamp": eval_data.get('timestamp', ''),
                        "video_url": eval_data.get('video_url', ''),
                        "video_title": eval_data.get('video_metadata', {}).get('title', '') or eval_data.get('title', ''),
                        "overall_score": eval_data.get('evaluation', {}).get('overall_score', 0.0),
                        "is_batch": is_batch,
                        "total_videos": total_videos
                    })
                except:
                    continue
        
        # 按时间戳倒序排序
        evaluations.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
        
        return jsonify({
            "success": True,
            "evaluations": evaluations
        })
    except Exception as e:
        logger.error(f"获取评估历史失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e),
            "evaluations": []
        }), 500

@app.route('/api/evaluation_reports', methods=['GET'])
def get_evaluation_reports():
    """获取评估报告列表（用于报告页面）"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        evaluations_dir = os.path.join(os.path.dirname(__file__), 'data', 'evaluations')
        if not os.path.exists(evaluations_dir):
            return jsonify({
                "success": True,
                "reports": [],
                "total_count": 0,
                "average_score": 0,
                "high_score_count": 0,
                "pending_count": 0
            })

        reports = []
        total_score = 0
        high_score_count = 0

        for filename in os.listdir(evaluations_dir):
            if filename.startswith('evaluation_') and filename.endswith('.json'):
                filepath = os.path.join(evaluations_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        eval_data = json.load(f)

                    # 跳过批量评估的主记录
                    if eval_data.get('is_batch', False):
                        continue

                    evaluation = eval_data.get('evaluation', {})
                    overall_score = evaluation.get('overall_score', 0)
                    metadata = eval_data.get('video_metadata', {})

                    # 获取搜索参数
                    search_params = eval_data.get('search_params', {})

                    reports.append({
                        "video_url": eval_data.get('video_url', ''),
                        "video_title": metadata.get('title', '') or eval_data.get('title', ''),
                        "total_score": overall_score,
                        "country": search_params.get('country', ''),
                        "grade": search_params.get('grade', ''),
                        "subject": search_params.get('subject', ''),
                        "evaluation_time": eval_data.get('timestamp', ''),
                        "ai_analysis": eval_data.get('analysis', '')
                    })

                    total_score += overall_score
                    if overall_score >= 8:
                        high_score_count += 1
                except:
                    continue

        # 按评估时间倒序排序
        reports.sort(key=lambda x: x.get('evaluation_time', ''), reverse=True)

        avg_score = total_score / len(reports) if reports else 0

        return jsonify({
            "success": True,
            "reports": reports,
            "total_count": len(reports),
            "average_score": avg_score,
            "high_score_count": high_score_count,
            "pending_count": 0  # 可以从搜索历史中计算未评估的视频数
        })
    except Exception as e:
        logger.error(f"获取评估报告失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@app.route('/api/evaluation_detail/<request_id>', methods=['GET'])
def get_evaluation_detail(request_id: str):
    """获取评估详情"""
    request_id_var = str(uuid.uuid4())[:8]
    set_request_id(request_id_var)
    
    try:
        evaluations_dir = os.path.join(os.path.dirname(__file__), 'data', 'evaluations')
        filepath = os.path.join(evaluations_dir, f"evaluation_{request_id}.json")
        
        if not os.path.exists(filepath):
            return jsonify({
                "success": False,
                "message": f"评估记录不存在: {request_id}"
            }), 404
        
        with open(filepath, 'r', encoding='utf-8') as f:
            eval_data = json.load(f)
        
        return jsonify({
            "success": True,
            "evaluation": eval_data
        })
    except Exception as e:
        logger.error(f"获取评估详情失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@app.route('/api/debug_logs', methods=['GET'])
def get_debug_logs():
    """获取Debug日志"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)
    
    try:
        lines = int(request.args.get('lines', 1000))
        since = request.args.get('since', '').strip()
        level = request.args.get('level', '').strip().upper()
        
        log_file = os.path.join(os.path.dirname(__file__), 'search_system.log')
        if not os.path.exists(log_file):
            return jsonify({
                "success": True,
                "logs": [],
                "total_lines": 0,
                "returned_lines": 0
            })
        
        with open(log_file, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()
        
        # 解析日志行
        parsed_logs = []
        for line in all_lines:
            line = line.strip()
            if not line:
                continue
            
            # 解析日志格式: 2025-12-29 15:00:54 UTC - logger_name - LEVEL - message
            parts = line.split(' - ', 3)
            if len(parts) >= 4:
                timestamp_str = parts[0]
                logger_name = parts[1]
                level_str = parts[2].upper()
                message = parts[3]
                
                # 解析时间戳
                try:
                    dt = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S UTC')
                    iso_timestamp = dt.replace(tzinfo=timezone.utc).isoformat()
                except:
                    iso_timestamp = timestamp_str
                
                # 时间过滤
                if since:
                    try:
                        since_dt = datetime.fromisoformat(since.replace('Z', '+00:00'))
                        if dt.replace(tzinfo=timezone.utc) < since_dt:
                            continue
                    except:
                        pass
                
                # 级别过滤
                if level and level_str != level:
                    continue
                
                parsed_logs.append({
                    "timestamp": timestamp_str,
                    "isoTimestamp": iso_timestamp,
                    "logger": logger_name,
                    "level": level_str.lower(),
                    "message": message
                })
        
        # 返回最后N行
        returned_logs = parsed_logs[-lines:] if len(parsed_logs) > lines else parsed_logs
        
        return jsonify({
            "success": True,
            "logs": returned_logs,
            "total_lines": len(parsed_logs),
            "returned_lines": len(returned_logs)
        })
    except Exception as e:
        logger.error(f"获取Debug日志失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e),
            "logs": []
        }), 500

@app.route('/api/save_debug_log', methods=['POST'])
def save_debug_log():
    """保存Debug日志到服务器"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)
    
    try:
        data = request.get_json()
        log_text = data.get('log_text', '')
        filename = data.get('filename', f"debug_log_{datetime.now().strftime('%Y-%m-%dT%H-%M-%S')}.txt")
        date_dir = data.get('date_dir', datetime.now().strftime('%Y-%m-%d'))
        
        if not log_text:
            return jsonify({
                "success": False,
                "message": "请提供日志内容"
            }), 400
        
        # 创建日期目录
        logs_dir = os.path.join(os.path.dirname(__file__), 'logs', date_dir)
        os.makedirs(logs_dir, exist_ok=True)
        
        # 保存文件
        filepath = os.path.join(logs_dir, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(log_text)
        
        logger.info(f"✅ 日志已保存: {filepath}")
        
        return jsonify({
            "success": True,
            "message": "日志已保存",
            "filepath": filepath
        })
    except Exception as e:
        logger.error(f"保存Debug日志失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@app.route('/api/export_excel', methods=['POST'])
def export_excel():
    """导出Excel"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        import io
        from datetime import datetime
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from config_manager import ConfigManager

        data = request.get_json()
        # 支持两种字段名：results 和 selected_results
        results = data.get('results') or data.get('selected_results', [])
        search_params = data.get('search_params', {})

        logger.info(f"[Excel导出] 收到数据: results={len(results)}个, search_params={search_params}")
        logger.info(f"[Excel导出] country={search_params.get('country')}, grade={search_params.get('grade')}, subject={search_params.get('subject')}")

        # 获取中文显示名称
        def get_chinese_display():
            """获取国家、年级、学科的中文显示"""
            country_code = search_params.get('country', '')
            grade_local = search_params.get('grade', '')
            subject_local = search_params.get('subject', '')

            # 优先使用前端传递的中文文本
            country_zh = search_params.get('countryText', country_code)
            grade_zh = search_params.get('gradeText', grade_local)
            subject_zh = search_params.get('subjectText', subject_local)

            # 如果前端没有提供中文文本，尝试从配置获取
            if country_zh == country_code or grade_zh == grade_local or subject_zh == subject_local:
                try:
                    config_manager = ConfigManager()
                    if country_code:
                        country_config = config_manager.get_country_config(country_code.upper())
                        if country_config:
                            # 只在需要时覆盖
                            if country_zh == country_code:
                                country_zh = country_config.country_name_zh or country_config.country_name

                            # 查找年级的中文名称
                            if grade_zh == grade_local:
                                for grade_info in country_config.grades:
                                    if grade_info['local_name'] == grade_local:
                                        grade_zh = grade_info['zh_name']
                                        break

                            # 查找学科的中文名称
                            if subject_zh == subject_local:
                                for subject_info in country_config.subjects:
                                    if subject_info['local_name'] == subject_local:
                                        subject_zh = subject_info['zh_name']
                                        break
                except Exception as e:
                    logger.warning(f"[Excel导出] 获取中文名称失败: {str(e)}")

            return country_zh, grade_zh, subject_zh

        country_zh, grade_zh, subject_zh = get_chinese_display()

        # 快速获取播放列表信息（视频数量和总时长）
        def get_playlist_info(url: str) -> tuple:
            """
            快速获取播放列表的视频数量和总时长

            Returns:
                (video_count, total_duration_minutes) - 如果失败返回 (None, None)
            """
            # 判断是否是播放列表URL
            if not url or 'list=' not in url:
                return None, None

            try:
                import yt_dlp

                ydl_opts = {
                    'quiet': True,
                    'no_warnings': True,
                    'extract_flat': True,  # 快速提取，不下载视频详情
                    'playlistend': None,  # 获取所有视频
                    'http_headers': {
                        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1',
                    },
                    'extractor_args': {
                        'youtube': {
                            'player_client': ['ios'],
                        }
                    },
                    'skip_download': True,
                }

                with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                    info = ydl.extract_info(url, download=False)

                if not info:
                    return None, None

                entries = info.get('entries', [])
                if not entries:
                    return None, None

                video_count = len(entries)

                # 计算总时长（分钟）
                total_duration_seconds = 0
                for entry in entries:
                    duration = entry.get('duration', 0)
                    if duration:
                        total_duration_seconds += duration

                total_duration_minutes = total_duration_seconds / 60 if total_duration_seconds > 0 else 0

                logger.info(f"[播放列表] URL: {url[:50]}..., 视频数: {video_count}, 总时长: {total_duration_minutes:.1f}分钟")

                return video_count, total_duration_minutes

            except Exception as e:
                logger.warning(f"[播放列表] 获取信息失败: {str(e)[:100]}")
                return None, None

        # 创建Excel数据
        excel_data = []
        for idx, r in enumerate(results, 1):
            # 获取分数、推荐理由和资源类型
            score = r.get('score', 0)
            recommendation_reason = r.get('recommendation_reason', r.get('recommendationReason', ''))
            resource_type = r.get('resource_type', r.get('resourceType', '未知'))
            url = r.get('url', '')

            # 获取播放列表信息
            video_count, total_duration = get_playlist_info(url)

            excel_data.append({
                '序号': idx,
                '国家': country_zh,
                '年级': grade_zh,
                '学科': subject_zh,
                '标题': r.get('title', ''),
                'URL': url,
                '摘要': r.get('snippet', '')[:500],  # 限制长度
                '资源类型': resource_type,
                '质量分数': score,
                '推荐理由': recommendation_reason,
                '来源': r.get('source', ''),
                '视频数量': video_count if video_count is not None else '-',
                '总时长(分钟)': f"{total_duration:.1f}" if total_duration and total_duration > 0 else '-',
            })

        # 创建DataFrame
        df = pd.DataFrame(excel_data)

        # 创建Excel writer
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='搜索结果', index=False)

            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['搜索结果']

            # 定义样式
            header_font = Font(bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(vertical='top', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 设置列宽
            column_widths = {
                'A': 6,   # 序号
                'B': 12,  # 国家
                'C': 12,  # 年级
                'D': 12,  # 学科
                'E': 50,  # 标题
                'F': 60,  # URL
                'G': 80,  # 摘要
                'H': 12,  # 资源类型
                'I': 10,  # 质量分数
                'J': 40,  # 推荐理由
                'K': 15,  # 来源
                'L': 12,  # 视频数量
                'M': 15,  # 总时长(分钟)
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # 应用表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 应用数据单元格样式
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # 设置行高
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                worksheet.row_dimensions[row[0].row].height = 60

        output.seek(0)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        country = search_params.get('country', 'Unknown')
        grade = search_params.get('grade', '')
        subject = search_params.get('subject', '')
        filename = f"{country}_{grade}_{subject}_{timestamp}.xlsx"

        logger.info(f"[Excel导出] 导出成功: {filename}, {len(results)} 行")

        # 返回文件
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except ImportError as e:
        logger.error(f"[Excel导出] 缺少依赖库: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"缺少必要的库，请安装: pip install pandas openpyxl"
        }), 500
    except Exception as e:
        logger.error(f"导出Excel失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

# ============================================================================
# 批量搜索Excel导出 API
# ============================================================================

@app.route('/api/export_batch_excel', methods=['POST'])
def export_batch_excel():
    """导出批量搜索结果到Excel"""
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        import io
        from datetime import datetime
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from config_manager import ConfigManager

        data = request.get_json()
        results = data.get('results', [])

        logger.info(f"[批量Excel导出] 开始导出 {len(results)} 个结果")

        # 创建Excel数据
        excel_data = []
        for idx, r in enumerate(results, 1):
            score = r.get('score', 0)
            recommendation_reason = r.get('recommendation_reason', r.get('recommendationReason', ''))
            resource_type = r.get('resource_type', r.get('resourceType', '未知'))
            url = r.get('url', '')

            # 获取批量搜索的国家、年级和学科信息
            # 兼容两种格式：batch_country/batch_grade/batch_subject 或 country/grade/subject
            batch_country = r.get('batch_country', r.get('country', ''))
            batch_grade = r.get('batch_grade', r.get('grade', ''))
            batch_subject = r.get('batch_subject', r.get('subject', ''))

            excel_data.append({
                '序号': idx,
                '国家': batch_country,
                '年级': batch_grade,
                '学科': batch_subject,
                '标题': r.get('title', ''),
                'URL': url,
                '摘要': r.get('snippet', '')[:500],
                '资源类型': resource_type,
                '质量分数': score,
                '推荐理由': recommendation_reason,
                '来源': r.get('source', ''),
                '视频数量': '-',  # 批量搜索时不获取视频数量（避免太慢）
                '总时长(分钟)': '-',
            })

        # 创建DataFrame
        df = pd.DataFrame(excel_data)

        # 创建Excel writer
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='批量搜索结果', index=False)

            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['批量搜索结果']

            # 定义样式
            header_font = Font(bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(vertical='top', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 设置列宽
            column_widths = {
                'A': 6,   # 序号
                'B': 12,  # 国家
                'C': 12,  # 年级
                'D': 12,  # 学科
                'E': 50,  # 标题
                'F': 60,  # URL
                'G': 80,  # 摘要
                'H': 12,  # 资源类型
                'I': 10,  # 质量分数
                'J': 40,  # 推荐理由
                'K': 15,  # 来源
                'L': 12,  # 视频数量
                'M': 15,  # 总时长(分钟)
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # 应用表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 应用数据单元格样式（分批处理，避免内存溢出）
            max_row = worksheet.max_row
            batch_size = 1000  # 每次处理1000行
            for start_row in range(2, max_row + 1, batch_size):
                end_row = min(start_row + batch_size - 1, max_row)
                for row in worksheet.iter_rows(min_row=start_row, max_row=end_row):
                    for cell in row:
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                # 每批处理后清理内存
                if start_row % (batch_size * 5) == 0:
                    import gc
                    gc.collect()

        # 生成文件名（日期(精确到分钟) + 国家 + 年级 + 学科）
        from collections import Counter

        # 获取所有唯一的国家、年级、学科
        countries = list(set([r.get('batch_country', '') for r in results]))
        grades = list(set([r.get('batch_grade', '') for r in results]))
        subjects = list(set([r.get('batch_subject', '') for r in results]))

        # 判断是否是全部年级/学科（通过检查数量）
        # 从ConfigManager获取配置来检查是否是全部
        config = ConfigManager()
        first_country_code = results[0].get('batch_country_code', '') if results else ''
        all_grades = []
        all_subjects = []

        if first_country_code:
            country_config = config.get_country_config(first_country_code)
            if country_config:
                # CountryConfig是Pydantic BaseModel，直接访问属性
                all_grades = [g.get('local_name', '') for g in country_config.grades]
                all_subjects = [s.get('local_name', '') for s in country_config.subjects]

        # 判断是否选择了全部
        is_all_grades = len(grades) >= len(all_grades) * 0.8 if all_grades else len(grades) > 1
        is_all_subjects = len(subjects) >= len(all_subjects) * 0.8 if all_subjects else len(subjects) > 1

        # 构建文件名组件
        country_part = countries[0] if len(countries) == 1 else f"{len(countries)}个国家"

        if is_all_grades:
            grade_part = "全部年级"
        else:
            # 提取年级名称（去掉中文括号部分）
            grade_part = "_".join([g.split(' (')[0] if ' (' in g else g for g in grades[:3]])
            if len(grades) > 3:
                grade_part += f"等{len(grades)}个"

        if is_all_subjects:
            subject_part = "全部学科"
        else:
            # 提取学科名称（去掉中文括号部分）
            subject_part = "_".join([s.split(' (')[0] if ' (' in s else s for s in subjects[:3]])
            if len(subjects) > 3:
                subject_part += f"等{len(subjects)}个"

        # 生成文件名：日期_国家_年级_学科
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        # 清理文件名中的特殊字符
        def clean_name(name):
            return name.replace(' ', '_').replace('/', '_').replace('\\', '_')[:20]

        filename = f"{timestamp}_{clean_name(country_part)}_{clean_name(grade_part)}_{clean_name(subject_part)}.xlsx"

        logger.info(f"[批量Excel导出] 文件名: {filename}")

        # 返回Excel文件（优化内存使用）
        output.seek(0)
        excel_bytes = output.getvalue()
        
        # 清理临时变量
        del excel_data, df, output
        import gc
        gc.collect()
        
        logger.info(f"[批量Excel导出] Excel文件大小: {len(excel_bytes) / 1024 / 1024:.2f} MB")
        
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except ImportError as e:
        logger.error(f"[批量Excel导出] 缺少依赖库: {str(e)}")
        return jsonify({
            "success": False,
            "message": f"缺少必要的库，请安装: pip install pandas openpyxl"
        }), 500
    except Exception as e:
        logger.error(f"批量导出Excel失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/api/export_search_log/<search_id>', methods=['GET'])
def export_search_log(search_id):
    """
    导出搜索日志为Excel文件

    Args:
        search_id: 搜索ID

    Returns:
        Excel文件下载
    """
    try:
        from core.search_log_collector import get_log_collector
        from core.excel_exporter import ExcelExporter

        logger.info(f"[日志导出] 开始导出搜索日志: {search_id}")

        # 获取日志收集器
        collector = get_log_collector()
        search_log = collector.get_log_by_id(search_id)

        if not search_log:
            logger.warning(f"[日志导出] 未找到搜索日志: {search_id}")
            return jsonify({
                "success": False,
                "error": f"未找到搜索日志: {search_id}"
            }), 404

        # 创建Excel导出器
        exporter = ExcelExporter()

        # 生成输出文件名
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"search_log_{search_id}_{timestamp}.xlsx"
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'logs',
            filename
        )

        # 确保logs目录存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # 导出Excel
        success = exporter.export_search_log(search_log, output_path)

        if success:
            logger.info(f"[日志导出] Excel文件已生成: {output_path}")
            # 发送文件
            return send_file(
                output_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        else:
            logger.error(f"[日志导出] 导出失败: {search_id}")
            return jsonify({
                "success": False,
                "error": "导出失败"
            }), 500

    except Exception as e:
        logger.error(f"[日志导出] 导出异常: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.route('/api/available_subjects', methods=['GET'])
def get_available_subjects():
    """
    获取指定年级的可用学科列表
    集成年级-学科联动规则

    Query Parameters:
        country: 国家代码（如：ID, CN, US）
        grade: 年级（如：Kelas 1, Grade 10, 一年级）

    Returns:
        {
            "success": true,
            "country": "ID",
            "grade": "Kelas 1",
            "subjects": [
                {
                    "local_name": "Matematika",
                    "zh_name": "数学",
                    "is_core": true,
                    "is_allowed": true,
                    "reason": ""
                },
                ...
            ]
        }
    """
    try:
        country = request.args.get('country', '').strip()
        grade = request.args.get('grade', '').strip()

        if not country or not grade:
            return jsonify({
                "success": False,
                "message": "请提供国家代码和年级",
                "subjects": []
            }), 400

        # 获取国家配置
        config = config_manager.get_country_config(country.upper())
        if not config:
            return jsonify({
                "success": False,
                "message": f"国家配置不存在: {country}",
                "subjects": []
            }), 404

        # 使用验证器获取可用学科
        validator = GradeSubjectValidator()
        subjects = config.subjects  # 获取所有学科
        available_subjects = validator.get_available_subjects(
            country.upper(),
            grade,
            subjects
        )

        # 只返回允许的学科
        allowed_subjects = [s for s in available_subjects if s.get('is_allowed', True)]

        return jsonify({
            "success": True,
            "country": country.upper(),
            "grade": grade,
            "subjects": allowed_subjects,
            "total_count": len(allowed_subjects)
        })

    except Exception as e:
        logger.error(f"获取可用学科失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e),
            "subjects": []
        }), 500


# ============================================================================
# 人工审核管理 API
# ============================================================================

@app.route('/api/review/submit', methods=['POST'])
def submit_review():
    """
    提交配置供人工审核

    Request Body:
        {
            "country_code": "ID",
            "country_name": "Indonesia",
            "changes": {...},
            "submitter": "admin",
            "reason": "自动生成的配对数据"
        }

    Returns:
        {
            "success": true,
            "review_id": "abc123",
            "message": "已提交审核"
        }
    """
    try:
        data = request.get_json()

        country_code = data.get('country_code', '').strip()
        country_name = data.get('country_name', '').strip()
        changes = data.get('changes', {})
        submitter = data.get('submitter', 'system').strip()
        reason = data.get('reason', '').strip()

        if not country_code or not country_name or not changes:
            return jsonify({
                "success": False,
                "message": "请提供国家代码、国家名称和变更内容"
            }), 400

        # 提交审核
        review_id = review_system.submit_for_review(
            country_code=country_code,
            country_name=country_name,
            changes=changes,
            submitter=submitter,
            reason=reason
        )

        return jsonify({
            "success": True,
            "review_id": review_id,
            "message": "已提交审核"
        })

    except Exception as e:
        logger.error(f"提交审核失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/api/review/list', methods=['GET'])
def list_reviews():
    """
    列出审核请求

    Query Parameters:
        status: 审核状态（可选）：pending, approved, rejected, changes_requested
        country_code: 国家代码（可选）

    Returns:
        {
            "success": true,
            "reviews": [...],
            "total_count": 10
        }
    """
    try:
        status_str = request.args.get('status', '').strip()
        country_code = request.args.get('country_code', '').strip()

        # 转换状态
        status = None
        if status_str:
            try:
                status = ReviewStatus(status_str)
            except ValueError:
                return jsonify({
                    "success": False,
                    "message": f"无效的状态值: {status_str}"
                }), 400

        # 获取审核列表
        reviews = review_system.list_review_requests(
            status=status,
            country_code=country_code if country_code else None
        )

        return jsonify({
            "success": True,
            "reviews": [r.model_dump() for r in reviews],
            "total_count": len(reviews)
        })

    except Exception as e:
        logger.error(f"获取审核列表失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e),
            "reviews": [],
            "total_count": 0
        }), 500


@app.route('/api/review/<review_id>', methods=['GET'])
def get_review(review_id):
    """
    获取单个审核请求的详细信息

    Returns:
        {
            "success": true,
            "review": {...}
        }
    """
    try:
        review = review_system.get_review_request(review_id)

        if not review:
            return jsonify({
                "success": False,
                "message": f"审核请求不存在: {review_id}"
            }), 404

        return jsonify({
            "success": True,
            "review": review.model_dump()
        })

    except Exception as e:
        logger.error(f"获取审核请求失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/api/review/approve', methods=['POST'])
def approve_review():
    """
    审核通过

    Request Body:
        {
            "review_id": "abc123",
            "reviewer": "admin",
            "comments": "数据准确"
        }

    Returns:
        {
            "success": true,
            "message": "审核通过"
        }
    """
    try:
        data = request.get_json()

        review_id = data.get('review_id', '').strip()
        reviewer = data.get('reviewer', 'admin').strip()
        comments = data.get('comments', '').strip()

        if not review_id:
            return jsonify({
                "success": False,
                "message": "请提供审核ID"
            }), 400

        # 审核通过
        success = review_system.approve_review(
            review_id=review_id,
            reviewer=reviewer,
            comments=comments
        )

        if not success:
            return jsonify({
                "success": False,
                "message": "审核通过失败，请检查审核ID"
            }), 404

        return jsonify({
            "success": True,
            "message": "审核通过"
        })

    except Exception as e:
        logger.error(f"审核通过失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/api/review/reject', methods=['POST'])
def reject_review():
    """
    审核拒绝

    Request Body:
        {
            "review_id": "abc123",
            "reviewer": "admin",
            "reason": "数据不准确"
        }

    Returns:
        {
            "success": true,
            "message": "审核拒绝"
        }
    """
    try:
        data = request.get_json()

        review_id = data.get('review_id', '').strip()
        reviewer = data.get('reviewer', 'admin').strip()
        reason = data.get('reason', '').strip()

        if not review_id or not reason:
            return jsonify({
                "success": False,
                "message": "请提供审核ID和拒绝原因"
            }), 400

        # 审核拒绝
        success = review_system.reject_review(
            review_id=review_id,
            reviewer=reviewer,
            reason=reason
        )

        if not success:
            return jsonify({
                "success": False,
                "message": "审核拒绝失败，请检查审核ID"
            }), 404

        return jsonify({
            "success": True,
            "message": "审核拒绝"
        })

    except Exception as e:
        logger.error(f"审核拒绝失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/api/review/request_changes', methods=['POST'])
def request_review_changes():
    """
    请求修改

    Request Body:
        {
            "review_id": "abc123",
            "reviewer": "admin",
            "comments": "需要补充XX信息"
        }

    Returns:
        {
            "success": true,
            "message": "已请求修改"
        }
    """
    try:
        data = request.get_json()

        review_id = data.get('review_id', '').strip()
        reviewer = data.get('reviewer', 'admin').strip()
        comments = data.get('comments', '').strip()

        if not review_id or not comments:
            return jsonify({
                "success": False,
                "message": "请提供审核ID和修改意见"
            }), 400

        # 请求修改
        success = review_system.request_changes(
            review_id=review_id,
            reviewer=reviewer,
            comments=comments
        )

        if not success:
            return jsonify({
                "success": False,
                "message": "请求修改失败，请检查审核ID"
            }), 404

        return jsonify({
            "success": True,
            "message": "已请求修改"
        })

    except Exception as e:
        logger.error(f"请求修改失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/api/review/statistics', methods=['GET'])
def get_review_statistics():
    """
    获取审核统计信息

    Returns:
        {
            "success": true,
            "statistics": {
                "total_reviews": 100,
                "pending_reviews": 10,
                "approved_reviews": 80,
                "rejected_reviews": 5,
                "changes_requested_reviews": 5
            }
        }
    """
    try:
        stats = review_system.get_statistics()

        return jsonify({
            "success": True,
            "statistics": stats.model_dump()
        })

    except Exception as e:
        logger.error(f"获取审核统计失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e),
            "statistics": {
                "total_reviews": 0,
                "pending_reviews": 0,
                "approved_reviews": 0,
                "rejected_reviews": 0,
                "changes_requested_reviews": 0
            }
        }), 500


# ============================================================================
# 大学教育搜索 API
# ============================================================================

@app.route('/api/feedback', methods=['POST'])
def submit_feedback():
    """
    提交用户反馈

    Request Body:
        {
            "result_id": "video_id",
            "search_params": {"country": "ID", "grade": "Kelas 1", "subject": "Matematika"},
            "explicit_feedback": {
                "rating": 5,
                "is_relevant": true,
                "text": "非常有帮助"
            },
            "implicit_signals": {
                "clicked": true,
                "dwell_time": 120,
                "scroll_depth": 0.8
            }
        }

    Returns:
        {
            "success": true,
            "feedback_id": "fb_20260108123456_abc12345"
        }
    """
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        if feedback_collector is None:
            return jsonify({
                "success": False,
                "message": "反馈收集系统未启用"
            }), 503

        data = request.get_json()

        result_id = data.get('result_id', '')
        search_params = data.get('search_params', {})
        explicit_feedback = data.get('explicit_feedback', {})
        implicit_signals = data.get('implicit_signals', {})
        system_context = data.get('system_context', {})

        # 收集反馈
        feedback_record = feedback_collector.collect_feedback(
            result_id=result_id,
            search_params=search_params,
            explicit_feedback=explicit_feedback,
            implicit_signals=implicit_signals,
            system_context=system_context
        )

        logger.info(f"✅ 反馈已收集: {feedback_record['feedback_id']}")

        return jsonify({
            "success": True,
            "feedback_id": feedback_record['feedback_id']
        })

    except Exception as e:
        logger.error(f"提交反馈失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/api/feedback/stats', methods=['GET'])
def get_feedback_stats():
    """
    获取反馈统计

    Query Parameters:
        days: 统计最近N天（默认7）

    Returns:
        {
            "success": true,
            "stats": {
                "period_days": 7,
                "total_feedbacks": 100,
                "avg_rating": 4.2,
                "relevance_rate": 0.85,
                "click_rate": 0.70,
                "rating_distribution": {...}
            }
        }
    """
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        if feedback_collector is None:
            return jsonify({
                "success": False,
                "message": "反馈收集系统未启用"
            }), 503

        days = int(request.args.get('days', 7))

        stats = feedback_collector.get_feedback_stats(days=days)

        return jsonify({
            "success": True,
            "stats": stats
        })

    except Exception as e:
        logger.error(f"获取反馈统计失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/api/admin/quality_evaluation', methods=['POST'])
@require_api_key  # ✅ 安全修复：需要API密钥认证
@require_admin  # ✅ 安全修复：需要管理员权限
def evaluate_quality():
    """
    评估搜索质量

    Request Body:
        {
            "results": [...],
            "search_params": {...}
        }

    Returns:
        {
            "success": true,
            "evaluation": {
                "overall_quality_score": 72.5,
                "quality_level": "良好",
                "basic_stats": {...},
                "anomalies": [...],
                "optimization_suggestions": [...]
            }
        }
    """
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        if quality_evaluator is None:
            return jsonify({
                "success": False,
                "message": "质量评估系统未启用"
            }), 503

        data = request.get_json()
        results = data.get('results', [])
        search_params = data.get('search_params', {})

        # 评估质量
        evaluation = quality_evaluator.evaluate_single_search(results, search_params)

        return jsonify({
            "success": True,
            "evaluation": evaluation
        })

    except Exception as e:
        logger.error(f"质量评估失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


def get_monitoring_dashboard():
    """
    获取监控仪表板数据

    Returns:
        {
            "success": true,
            "dashboard": {
                "system_health": {...},
                "recent_alerts": [...],
                "metrics_trend": {...},
                "current_metrics": {...}
            }
        }
    """
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        if monitoring_system is None:
            return jsonify({
                "success": False,
                "message": "监控系统未启用"
            }), 503

        dashboard_data = monitoring_system.generate_dashboard_data()

        return jsonify({
            "success": True,
            "dashboard": dashboard_data
        })

    except Exception as e:
        logger.error(f"获取监控仪表板失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/api/admin/system_health', methods=['GET'])
def get_system_health():
    """
    获取系统健康状态

    Returns:
        {
            "success": true,
            "health": {
                "health_score": 85,
                "health_level": "健康",
                "issues": [],
                "current_metrics": {...}
            }
        }
    """
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        if monitoring_system is None:
            return jsonify({
                "success": False,
                "message": "监控系统未启用"
            }), 503

        health = monitoring_system.get_system_health()

        return jsonify({
            "success": True,
            "health": health
        })

    except Exception as e:
        logger.error(f"获取系统健康状态失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


def get_optimization_status():
    """
    获取优化系统状态

    Returns:
        {
            "success": true,
            "status": {
                "environment": "staging",
                "feature_flags": {...},
                "optimization_history": [...],
                "last_run": "2026-01-08T12:00:00Z"
            }
        }
    """
    request_id = str(uuid.uuid4())[:8]
    set_request_id(request_id)

    try:
        import yaml

        # 读取功能开关配置
        config_path = os.path.join(os.path.dirname(__file__), 'config', 'feature_flags.yaml')
        with open(config_path, 'r', encoding='utf-8') as f:
            feature_flags = yaml.safe_load(f)

        return jsonify({
            "success": True,
            "status": {
                "environment": sis_environment,
                "feature_flags": feature_flags.get('feature_flags', {}),
                "monitoring_enabled": monitoring_system is not None,
                "feedback_enabled": feedback_collector is not None,
                "quality_evaluation_enabled": quality_evaluator is not None
            }
        })

    except Exception as e:
        logger.error(f"获取优化状态失败: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


def sis_dashboard():
    """Self-Improving System管理仪表板"""
    return render_template('sis_dashboard.html')


# ============================================================================
# 智能优化审批 API
# ============================================================================

if __name__ == '__main__':
    # 从环境变量读取端口，默认5000，如果被占用则尝试5001, 5002等
    import os
    import socket
    
    def find_free_port(start_port=5000, max_attempts=10):
        """查找可用端口"""
        for i in range(max_attempts):
            port = start_port + i
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            try:
                sock.bind(('0.0.0.0', port))
                sock.close()
                return port
            except OSError:
                continue
        raise RuntimeError(f"无法找到可用端口（尝试了 {start_port}-{start_port+max_attempts-1}）")
    
    # 优先使用环境变量指定的端口
    env_port = os.getenv('FLASK_PORT')
    if env_port:
        try:
            port = int(env_port)
        except ValueError:
            logger.warning(f"无效的FLASK_PORT环境变量: {env_port}，将自动查找可用端口")
            port = find_free_port()
    else:
        # 自动查找可用端口
        port = find_free_port()
    
    logger.info(f"🌐 启动Web服务器，端口: {port}")
    print(f"\n{'='*60}")
    print(f"🚀 Web应用已启动")
    print(f"📌 访问地址: http://localhost:{port}")
    print(f"{'='*60}\n")
    
    app.run(debug=True, host='0.0.0.0', port=port, use_reloader=False)
