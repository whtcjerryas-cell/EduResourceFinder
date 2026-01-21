#!/usr/bin/env python3
"""
导出处理器 - 拆分 web_app.py 中的超长 export_excel() 函数
"""

import io
from datetime import datetime
from typing import Dict, Any, List, Tuple, Optional
from utils.logger_utils import get_logger

logger = get_logger('export_handler')


class ExportHandler:
    """导出处理器 - 封装Excel导出的完整流程"""

    def __init__(self, config_manager=None):
        """
        初始化导出处理器

        Args:
            config_manager: 配置管理器（可选）
        """
        if config_manager is None:
            from config_manager import ConfigManager
            config_manager = ConfigManager()
        self.config_manager = config_manager

    def export_search_results_to_excel(
        self,
        results: List[Dict[str, Any]],
        search_params: Dict[str, Any]
    ) -> Tuple[io.BytesIO, str]:
        """
        导出搜索结果到Excel

        Args:
            results: 搜索结果列表
            search_params: 搜索参数

        Returns:
            (Excel文件对象, 文件名)
        """
        # 1. 准备数据
        excel_data = self._prepare_excel_data(results, search_params)

        # 2. 生成Excel
        output = self._generate_excel(excel_data)

        # 3. 生成文件名
        filename = self._generate_filename(search_params)

        return output, filename

    def _prepare_excel_data(
        self,
        results: List[Dict[str, Any]],
        search_params: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        准备Excel数据

        Args:
            results: 搜索结果
            search_params: 搜索参数

        Returns:
            Excel数据列表
        """
        # 获取中文显示名称
        country_zh, grade_zh, subject_zh = self._get_chinese_display_names(search_params)

        excel_data = []
        for idx, r in enumerate(results, 1):
            score = r.get('score', 0)
            recommendation_reason = r.get('recommendation_reason', r.get('recommendationReason', ''))
            resource_type = r.get('resource_type', r.get('resourceType', '未知'))
            url = r.get('url', '')

            # 获取播放列表信息
            video_count, total_duration = self._get_playlist_info(url)

            excel_data.append({
                '序号': idx,
                '国家': country_zh,
                '年级': grade_zh,
                '学科': subject_zh,
                '标题': r.get('title', ''),
                'URL': url,
                '摘要': r.get('snippet', '')[:500],
                '资源类型': resource_type,
                '质量分数': score,
                '推荐理由': recommendation_reason,
                '来源': r.get('source', ''),
                '视频数量': video_count if video_count else '-',
                '总时长(分钟)': f"{total_duration:.1f}" if total_duration and total_duration > 0 else '-',
            })

        return excel_data

    def _get_chinese_display_names(self, search_params: Dict[str, Any]) -> Tuple[str, str, str]:
        """
        获取国家、年级、学科的中文显示名称

        Args:
            search_params: 搜索参数

        Returns:
            (country_zh, grade_zh, subject_zh)
        """
        from utils.helpers import get_chinese_display_names
        return get_chinese_display_names(self.config_manager, search_params)

    def _get_playlist_info(self, url: str) -> Tuple[Optional[int], Optional[float]]:
        """
        快速获取播放列表的视频数量和总时长

        Args:
            url: 播放列表URL

        Returns:
            (video_count, total_duration_minutes) - 如果失败返回 (None, None)
        """
        if not url or 'list=' not in url:
            return None, None

        try:
            import yt_dlp

            ydl_opts = {
                'quiet': True,
                'no_warnings': True,
                'extract_flat': True,
                'playlistend': None,
                'http_headers': {
                    'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1',
                },
                'extractor_args': {
                    'youtube': {
                        'player_client': ['ios'],
                    }
                },
                'skip_download': True,
            }

            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                info = ydl.extract_info(url, download=False)

            if not info:
                return None, None

            entries = info.get('entries', [])
            if not entries:
                return None, None

            video_count = len(entries)

            # 计算总时长（分钟）
            total_duration_seconds = 0
            for entry in entries:
                duration = entry.get('duration', 0)
                if duration:
                    total_duration_seconds += duration

            total_duration_minutes = total_duration_seconds / 60 if total_duration_seconds > 0 else 0

            logger.info(f"[播放列表] URL: {url[:50]}..., 视频数: {video_count}, 总时长: {total_duration_minutes:.1f}分钟")

            return video_count, total_duration_minutes

        except Exception as e:
            logger.warning(f"[播放列表] 获取信息失败: {str(e)[:100]}")
            return None, None

    def _generate_excel(self, excel_data: List[Dict[str, Any]]) -> io.BytesIO:
        """
        生成Excel文件

        Args:
            excel_data: Excel数据列表

        Returns:
            Excel文件对象（BytesIO）
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from utils.constants import (
                EXCEL_COLUMN_WIDTHS,
                EXCEL_HEADER_ROW_HEIGHT,
                EXCEL_DATA_ROW_HEIGHT
            )
        except ImportError as e:
            logger.error(f"[Excel导出] 缺少依赖库: {str(e)}")
            raise

        # 创建DataFrame
        df = pd.DataFrame(excel_data)

        # 创建Excel writer
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='搜索结果', index=False)

            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['搜索结果']

            # 定义样式
            header_font = Font(bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(vertical='top', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 设置列宽
            for col, width in EXCEL_COLUMN_WIDTHS.items():
                worksheet.column_dimensions[col].width = width

            # 应用表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                worksheet.row_dimensions[cell.row].height = EXCEL_HEADER_ROW_HEIGHT

            # 应用数据单元格样式
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                worksheet.row_dimensions[row[0].row].height = EXCEL_DATA_ROW_HEIGHT

        output.seek(0)
        return output

    def _generate_filename(self, search_params: Dict[str, Any]) -> str:
        """
        生成文件名

        Args:
            search_params: 搜索参数

        Returns:
            文件名
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        country = search_params.get('country', 'Unknown')
        grade = search_params.get('grade', '')
        subject = search_params.get('subject', '')
        return f"{country}_{grade}_{subject}_{timestamp}.xlsx"
